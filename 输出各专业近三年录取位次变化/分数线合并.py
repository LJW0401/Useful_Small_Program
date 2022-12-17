import os
import openpyxl

path_2022 = os.path.dirname(__file__) + '/浙江省2022年普通高校招生普通类第一段平行志愿.xlsx'
path_2021 = os.path.dirname(__file__) + '/浙江省2021年普通高校招生普通类第一段平行投档分数线.xlsx'
path_2020 = os.path.dirname(__file__) + '/浙江省2020年普通高校招生普通类第一段平行投档分数线.xlsx'
path_2019 = os.path.dirname(__file__) + '/浙江省2019年普通高校招生普通类第一段平行投档分数线.xlsx'

work_book_2022 = openpyxl.load_workbook(path_2022)
work_book_2021 = openpyxl.load_workbook(path_2021)
work_book_2020 = openpyxl.load_workbook(path_2020)
work_book_2019 = openpyxl.load_workbook(path_2019)


print(work_book_2022.sheetnames,work_book_2021.sheetnames)



