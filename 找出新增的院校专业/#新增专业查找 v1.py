import os
import datetime
import openpyxl
#注：本程序以院校代码为关键字判断是否为同一院校
#加载区分文件
const = "0ABCDEFGHIJKLMN"
#加载两年的专业信息
def LoadInfor(FileNow,SheetNow,FilePast,SheetPast,OutputFile='2022年新增院校专业',OutputSheet='新增院校专业'):
    global WorkSheetNow,WorkSheetPast,OutWorkBook,OutWorkSheet
    path1 = os.path.dirname(__file__) + f'/{FileNow}.xlsx'
    path2 = os.path.dirname(__file__) + f'/{FilePast}.xlsx'
    WorkBookNow = openpyxl.load_workbook(path1)
    WorkBookPast = openpyxl.load_workbook(path2)
    WorkSheetNow = WorkBookNow[SheetNow]
    WorkSheetPast = WorkBookPast[SheetPast]
    path3 = os.path.dirname(__file__) + f'/{OutputFile}.xlsx'
    OutWorkBook = openpyxl.load_workbook(path3)
    OutWorkSheet = OutWorkBook[OutputSheet]

def LockRow(sheet,code,RowMin,RowMax):#锁定院校代码所在的row的范围
    locked = False
    RowMin = RowMax
    while int(sheet.cell(row = RowMin,column = 1).value) < code:
        RowMin += 1
    if int(sheet.cell(row = RowMin,column = 1).value) == code:
        RowMax = RowMin + 1
        while sheet.cell(row = RowMax,column = 1).value == sheet.cell(row = RowMax - 1,column = 1).value:
            RowMax += 1
        RowMax -= 1
        locked = True
    else:
        RowMax = RowMin
    #print(RowMin,RowMax)
    return RowMin,RowMax,locked

def DelBrackets(txt = ''):#去除文本中的括号
    while '(' in txt and ')' in txt:
        txt = txt[:txt.find('(')] + txt[txt.find(')') + 1:]
    while '（' in txt and '）' in txt:
        txt = txt[:txt.find('（')] + txt[txt.find('）') + 1:]
    return txt

def check(subject,sheet,RowMin,RowMax,):#在去年院校范围内查询专业是否存在
    flag = False
    for row in range(RowMin,RowMax + 1):
        if sheet[f'D{row}'].value == subject:
            flag = True
            break
        elif DelBrackets(sheet[f'D{row}'].value) == DelBrackets(subject):
            flag = True
            break
    return flag

def FindDifference(sheet1,sheet2):#查找去年没有的专业
    NowCode = 1 #现在检测的院校的代码
    RowTemp = 1 #临时寄存当前查询单元格的row
    P_RowMin = 2 #表示去年志愿中该专业的row的最小值
    P_RowMax = 2 #表示去年志愿中该专业的row的最大值
    Locked = False #表示已经在去年志愿表格中锁定row的范围
    OutRow = 2 #记录输出文件中的row
    output = False#表示是否输出
    for line in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row, min_col=1, max_col=sheet1.max_column):
        output = False
        for cell in line:
            #今年检测院校代码与去年检测院校代码不一致
            if (cell.column == 'A' and int(cell.value) != NowCode) or (cell.column == 'A' and NowCode == 1 and not Locked):
                #s1取消锁定状态，s2在去年表格中锁定对应院校专业row范围，如无院校直接增加到输出文件
                #s3如无专业则输出到文件
                #step-1取消锁定状态
                Locked = False
                NowCode = int(cell.value)
                P_RowMin,P_RowMax,Locked = LockRow(sheet2,NowCode,P_RowMin,P_RowMax)
                print(f'正在比对 {DelBrackets(sheet1.cell(row = cell.row,column = 2).value)} 信息')
            if not Locked:#去年无该院校则输出
                OutWorkSheet[f'{cell.column}{OutRow}'].value = cell.value
                output = True
                RowTemp = cell.row
            elif cell.column == 'D' and not check(cell.value,sheet2,P_RowMin,P_RowMax):#检测专业是否存在
                output = True
                OutWorkSheet[f'{cell.column}{OutRow}'].value = cell.value
                RowTemp = cell.row
            elif cell.column > 'D' and output:
                OutWorkSheet[f'{cell.column}{OutRow}'].value = cell.value
        if output :
            if OutWorkSheet[f'A{OutRow}'].value != sheet1[f'A{RowTemp}']:
                OutWorkSheet[f'A{OutRow}'].value = sheet1[f'A{RowTemp}'].value
                OutWorkSheet[f'B{OutRow}'].value = sheet1[f'B{RowTemp}'].value
                OutWorkSheet[f'C{OutRow}'].value = sheet1[f'C{RowTemp}'].value
            OutRow += 1

#----------程序运行区----------
StartTime = datetime.datetime.now()
LoadInfor('浙江省2022年志愿填报','Sheet1','浙江省2021年普通高校招生普通类第一段平行投档分数线','tdx2021')
print('---已打开文件---')
for cell in WorkSheetNow["A1:L1"][0]:
    OutWorkSheet[f'{cell.column}1'].value = cell.value
FindDifference(WorkSheetNow,WorkSheetPast)
#----------保存输出文件----------
OutputFile = '2022年新增院校专业'
OutWorkBook.save(os.path.dirname(__file__) + f'/{OutputFile}.xlsx')
#----------程序运行计时----------
EndTime = datetime.datetime.now()
DeltaTime = EndTime - StartTime
print(f'\n-----运行耗时{DeltaTime}-----\n')